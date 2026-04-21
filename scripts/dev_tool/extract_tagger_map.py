#!/usr/bin/env python3
"""
extract_tagger_map.py — convert ``docs/Tagger_translation_0.xlsx`` into a
JSON channel map that the tagger TDC viewer (and anything else that wants
human-readable counter names) can consume.

Run:
    pip install openpyxl       # one-time (if not already present)
    python scripts/extract_tagger_map.py \
        [--in docs/Tagger_translation_0.xlsx] \
        [--out database/tagger_map.json]

The top block of the spreadsheet (rows 7-22) holds the V1190 TDC map used
by the tagger crate.  Each of slot 18, 19, 20 has a (channel, counter) pair
of columns — we read those pairs and drop entries that are blank or tagged
``not used``.  The scaler/discriminator block further down the sheet is
separate hardware and is not included here.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path


# Spreadsheet layout (1-based column indices).  Update these if a future
# revision of the xlsx rearranges columns.
SLOT_COLUMNS = [
    # (slot, channel_col, name_col)
    (18, 1, 2),
    (19, 4, 5),
    (20, 7, 8),
]
ROW_FIRST = 7
ROW_LAST = 22


def extract(xlsx_path: Path) -> list[dict]:
    import openpyxl  # imported here so the rest of the file is introspectable

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"]

    entries: list[dict] = []
    for slot, ch_col, name_col in SLOT_COLUMNS:
        for r in range(ROW_FIRST, ROW_LAST + 1):
            ch = ws.cell(row=r, column=ch_col).value
            name = ws.cell(row=r, column=name_col).value
            if ch is None:
                continue
            if isinstance(name, str):
                name = name.strip()
            if not name or (isinstance(name, str) and name.lower() in ("not used", "-")):
                continue
            entries.append({"slot": int(slot), "channel": int(ch), "name": str(name)})
    return entries


def main() -> int:
    here = Path(__file__).resolve().parent.parent
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--in", dest="inp", type=Path,
                    default=here / "docs" / "Tagger_translation_0.xlsx")
    ap.add_argument("--out", dest="outp", type=Path,
                    default=here / "database" / "tagger_map.json")
    args = ap.parse_args()

    if not args.inp.exists():
        ap.error(f"input not found: {args.inp}")

    entries = extract(args.inp)
    payload = {
        "_comment": ("Tagger V1190 TDC channel map - slot, channel -> counter/detector "
                     "name. Extracted from docs/Tagger_translation_0.xlsx."),
        "source": str(args.inp.name),
        "channels": entries,
    }
    args.outp.parent.mkdir(parents=True, exist_ok=True)
    with args.outp.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    print(f"wrote {len(entries)} entries to {args.outp}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
